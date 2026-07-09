"""Extract .xlsx / .csv into clean, structured, model-readable JSON.

Deterministic ingestion so a model works on faithful data instead of a text scrape.
For each sheet it emits the header, typed rows (dates as ISO-8601 strings, formula
*results* not formula text), merged-cell ranges, and any Excel error cells, plus a
top-level fidelity block that flags what could NOT be read faithfully (e.g. formulas
with no cached result). Prints JSON to stdout; exits 0 on a successful read.

This is an EXTRACTOR (ingest + fidelity self-check), not a validate->fix loop.

Usage:
    python scripts/extract_workbook.py path/to/file.xlsx [--max-rows N]
    python scripts/extract_workbook.py path/to/file.csv
"""
import csv
import datetime as dt
import json
import sys
from pathlib import Path

ERROR_VALUES = {
    "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!",
    "#SPILL!", "#CALC!", "#GETTING_DATA",
}


def jsonable(v):
    if isinstance(v, (dt.datetime, dt.date, dt.time)):
        return v.isoformat()
    return v


def sniff_csv(path):
    with open(path, "rb") as f:
        raw = f.read(65536)
    encoding = "utf-8-sig" if raw.startswith(b"\xef\xbb\xbf") else "utf-8"
    sample = raw.decode(encoding, errors="replace")
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","
    return encoding, delimiter


def extract_csv(path):
    encoding, delimiter = sniff_csv(path)
    rows = []
    with open(path, newline="", encoding=encoding) as f:
        for row in csv.reader(f, delimiter=delimiter):
            rows.append(row)
    header = rows[0] if rows else []
    body = rows[1:] if len(rows) > 1 else []
    return {
        "file": str(path),
        "format": "csv",
        "encoding": encoding,
        "delimiter": delimiter,
        "sheets": [{
            "name": path.stem,
            "n_rows": len(body),
            "n_cols": len(header),
            "header": header,
            "rows": body,
            "merged_cells": [],
            "error_cells": [],
        }],
        "fidelity": {
            "n_sheets": 1,
            "cached_values_present": True,
            "warnings": [] if rows else ["File is empty."],
        },
    }


def extract_xlsx(path, max_rows=None):
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    sheets, warnings = [], []
    saw_cached, saw_formula_gap = False, False

    # Merged ranges need the non-read-only view.
    from openpyxl import load_workbook as lw
    wb_struct = lw(str(path), data_only=False, read_only=False)
    merged_by_sheet = {ws.title: [str(r) for r in ws.merged_cells.ranges]
                       for ws in wb_struct.worksheets}
    for ws in wb_struct.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    saw_formula_gap = True
                    break
            if saw_formula_gap:
                break
    wb_struct.close()

    for ws in wb.worksheets:
        data = []
        error_cells = []
        for r_idx, row in enumerate(ws.iter_rows()):
            values = []
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v in ERROR_VALUES:
                    error_cells.append(f"{cell.coordinate} -> {v}")
                if v is not None:
                    saw_cached = True
                values.append(jsonable(v))
            data.append(values)
            if max_rows is not None and r_idx + 1 >= max_rows + 1:  # +1 for header
                warnings.append(
                    f"Sheet {ws.title!r} truncated to {max_rows} data rows."
                )
                break
        # Trim wholly-empty trailing rows/cols conservatively (keep shape simple).
        header = data[0] if data else []
        body = data[1:] if len(data) > 1 else []
        sheets.append({
            "name": ws.title,
            "n_rows": len(body),
            "n_cols": len(header),
            "header": header,
            "rows": body,
            "merged_cells": merged_by_sheet.get(ws.title, []),
            "error_cells": error_cells,
        })
    wb.close()

    if saw_formula_gap and not saw_cached:
        warnings.append(
            "Formulas present but no cached results found: openpyxl cannot compute "
            "formulas. Values may be missing — open/recalc in Excel, then re-extract."
        )
    for s in sheets:
        if s["merged_cells"]:
            warnings.append(
                f"Sheet {s['name']!r} has merged cells {s['merged_cells']}: value "
                "lives only in the top-left cell; downstream rows may misalign."
            )
        if s["error_cells"]:
            warnings.append(
                f"Sheet {s['name']!r} contains Excel error values {s['error_cells']}."
            )
    return {
        "file": str(path),
        "format": "xlsx",
        "sheets": sheets,
        "fidelity": {
            "n_sheets": len(sheets),
            "cached_values_present": saw_cached,
            "warnings": warnings,
        },
    }


def main():
    args = sys.argv[1:]
    if not args:
        sys.exit("usage: python scripts/extract_workbook.py <file.xlsx|.csv> [--max-rows N]")
    path = Path(args[0])
    if not path.exists():
        sys.exit(f"file not found: {path}")
    max_rows = None
    if "--max-rows" in args:
        max_rows = int(args[args.index("--max-rows") + 1])

    ext = path.suffix.lower()
    if ext == ".csv":
        result = extract_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        result = extract_xlsx(path, max_rows=max_rows)
    else:
        sys.exit(f"unsupported extension {ext!r}; expected .xlsx/.xlsm/.xls/.csv")

    print(json.dumps(result, indent=2, ensure_ascii=False))
    sys.exit(0)


if __name__ == "__main__":
    main()

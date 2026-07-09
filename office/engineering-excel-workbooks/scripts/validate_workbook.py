"""Validate an .xlsx workbook before it ships.

Deterministic quality gate for a produced workbook. Checks the failure modes that
actually matter for Excel and prints a machine-readable JSON report so the caller
can read the errors, fix them, and re-run until clean (the produce -> validate ->
fix -> re-validate loop). Exits non-zero when there are ERRORS.

Checks:
  ERROR   - Excel error values in cached results (#REF!, #DIV/0!, #VALUE!, #NAME?,
            #N/A, #NULL!, #NUM!, #SPILL!, #CALC!, #GETTING_DATA).
  ERROR   - Leftover placeholder text (TBD, TODO, XXX, lorem/ipsum, FIXME,
            PLACEHOLDER, and {{ }} / {% %} template tags).
  WARNING - Merged cells (break sorting, filtering, and references).
  WARNING - Empty worksheets.
  WARNING - External workbook links (fragile cross-file references).
  WARNING - No cached formula values found (openpyxl cannot compute formulas;
            Excel error results cannot be checked until the file is recalculated).

Usage:
    python scripts/validate_workbook.py path/to/workbook.xlsx
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ERROR_VALUES = {
    "#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!",
    "#SPILL!", "#CALC!", "#GETTING_DATA",
}
PLACEHOLDER = re.compile(
    r"\bTBD\b|\bTODO\b|\bFIXME\b|\bXXX+\b|\bPLACEHOLDER\b|lorem\s+ipsum|{{.*?}}|{%.*?%}",
    re.IGNORECASE,
)


def scan_values(path):
    """Return (error_hits, placeholder_hits, saw_cached_value, sheet_names).

    Loads with data_only=True so cached formula results (including error strings)
    are read as their computed values.
    """
    wb = load_workbook(str(path), data_only=True, read_only=True)
    error_hits, placeholder_hits = [], []
    saw_cached_value = False
    empty_sheets = []
    for ws in wb.worksheets:
        non_empty = False
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                non_empty = True
                if isinstance(v, str):
                    if v in ERROR_VALUES:
                        error_hits.append(f"{ws.title}!{cell.coordinate} -> {v}")
                    m = PLACEHOLDER.search(v)
                    if m:
                        placeholder_hits.append(
                            f"{ws.title}!{cell.coordinate} -> {m.group(0)!r}"
                        )
                    saw_cached_value = True
                else:
                    saw_cached_value = True
        if not non_empty:
            empty_sheets.append(ws.title)
    wb.close()
    return error_hits, placeholder_hits, saw_cached_value, empty_sheets, wb.sheetnames


def scan_structure(path):
    """Return (merged, has_formulas, external_links) using the formula view."""
    wb = load_workbook(str(path), data_only=False, read_only=False)
    merged, has_formulas = [], False
    for ws in wb.worksheets:
        for rng in ws.merged_cells.ranges:
            merged.append(f"{ws.title}!{rng}")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    has_formulas = True
                    break
            if has_formulas:
                break
    external = []
    try:
        external = [str(link.file_link.Target) for link in wb._external_links]
    except Exception:
        external = []
    wb.close()
    return merged, has_formulas, external


def main():
    if len(sys.argv) != 2:
        sys.exit("usage: python scripts/validate_workbook.py <file.xlsx>")
    path = Path(sys.argv[1])
    if not path.exists():
        sys.exit(f"file not found: {path}")

    err_vals, placeholders, saw_cached, empty_sheets, sheets = scan_values(path)
    merged, has_formulas, external = scan_structure(path)

    errors, warnings = [], []
    if err_vals:
        errors.append(f"Excel error values in cells: {err_vals}")
    if placeholders:
        errors.append(f"Leftover placeholder text: {placeholders}")

    if merged:
        warnings.append(f"Merged cells present (break sort/filter/refs): {merged}")
    if empty_sheets:
        warnings.append(f"Empty worksheets: {empty_sheets}")
    if external:
        warnings.append(f"External workbook links (fragile): {external}")
    if has_formulas and not saw_cached:
        warnings.append(
            "Formulas exist but no cached results were found. openpyxl does not "
            "compute formulas; open and save in Excel (or run a recalc step) so "
            "error results can be checked, then re-run this validator."
        )

    report = {
        "file": str(path),
        "status": "OK" if not errors else "FAILED",
        "errors": errors,
        "warnings": warnings,
        "sheets": sheets,
    }
    print(json.dumps(report, indent=2))
    sys.exit(0 if not errors else 1)


if __name__ == "__main__":
    main()
